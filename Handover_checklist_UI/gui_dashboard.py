import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import json
import os
import sys
import subprocess
from openpyxl import Workbook

# Add the parent directory to the Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import functions from other scripts using the correct relative paths
from File_Processing.get_all_files_register import get_all_files
from File_Checking.check_file_names import categorize_files

def process_directory(directory):
    file_list = get_all_files(directory)
    return categorize_files(file_list)

class HandoverChecklist(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Handover Checklist Dashboard")
        self.geometry("1000x600")

        self.folder_path = tk.StringVar()
        self.processed_data = None
        self.create_widgets()

    def create_widgets(self):
        # Folder selection
        tk.Label(self, text="Select Directory:").pack(pady=5)
        tk.Entry(self, textvariable=self.folder_path, width=50).pack(side=tk.LEFT, padx=5)
        tk.Button(self, text="Browse", command=self.browse_folder).pack(side=tk.LEFT)
        tk.Button(self, text="Process Directory", command=self.process_directory).pack(pady=5)

        # Excel generation button
        self.excel_button = tk.Button(self, text="Generate Excel Register", command=self.generate_excel, state=tk.DISABLED)
        self.excel_button.pack(pady=5)

        # Table
        self.tree = ttk.Treeview(self, columns=("Category", "File Name", "Full Path"), show="tree headings")
        self.tree.heading("Category", text="Category")
        self.tree.heading("File Name", text="File Name")
        self.tree.heading("Full Path", text="Full Path")
        self.tree.column("Category", width=200)
        self.tree.column("File Name", width=300)
        self.tree.column("Full Path", width=500)
        self.tree.pack(fill=tk.BOTH, expand=1)

        # Vertical Scrollbar
        scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Horizontal Scrollbar
        h_scrollbar = ttk.Scrollbar(self, orient=tk.HORIZONTAL, command=self.tree.xview)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.configure(xscroll=h_scrollbar.set)

        # Bind double-click event
        self.tree.bind("<Double-1>", self.on_double_click)

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        self.folder_path.set(folder_selected)

    def process_directory(self):
        folder = self.folder_path.get()
        if folder:
            self.processed_data = process_directory(folder)
            self.update_table(self.processed_data)
            self.excel_button.config(state=tk.NORMAL)  # Enable Excel button
        else:
            tk.messagebox.showwarning("Warning", "Please select a directory first!")

    def update_table(self, data):
        # Clear existing data
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        # Insert new data
        for category, files in data.items():
            category_id = self.tree.insert("", tk.END, text=category, values=(category, f"{len(files)} files found"), tags=('category',))
            for file in files:
                file_name = os.path.basename(file)
                self.tree.insert(category_id, tk.END, values=(category, file_name, file), tags=('file',))
        
        # Configure tag colors
        self.tree.tag_configure('category', background='lightgrey')
        self.tree.tag_configure('file', background='white')

    def on_double_click(self, event):
        item = self.tree.identify('item', event.x, event.y)
        if not item:
            return  # Clicked on empty space

        values = self.tree.item(item, 'values')
        tags = self.tree.item(item, 'tags')

        if 'category' in tags:
            # Clicked on a category row, do nothing or toggle expand/collapse
            return

        if 'file' in tags and len(values) >= 3:
            file_path = values[2]  # Full path should be in the third column
            if os.path.exists(file_path):
                self.open_file(file_path)
            else:
                messagebox.showerror("Error", f"File not found: {file_path}")
        else:
            messagebox.showerror("Error", "Invalid file selection")

    def open_file(self, file_path):
        try:
            if sys.platform == "win32":
                os.startfile(file_path)
            elif sys.platform == "darwin":
                subprocess.call(["open", file_path])
            else:
                subprocess.call(["xdg-open", file_path])
        except Exception as e:
            tk.messagebox.showerror("Error", f"Unable to open file: {str(e)}")

    def generate_excel(self):
        if not self.processed_data:
            tk.messagebox.showwarning("Warning", "Please process a directory first!")
            return

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Handover Checklist"

        # Write headers
        ws['A1'] = "Category"
        ws['B1'] = "File Path"

        # Write data
        row = 2
        for category, files in self.processed_data.items():
            for file in files:
                ws.cell(row=row, column=1, value=category)
                ws.cell(row=row, column=2, value=file)
                row += 1

        # Save the workbook
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            wb.save(save_path)
            tk.messagebox.showinfo("Success", f"Excel file saved successfully at {save_path}")
        else:
            tk.messagebox.showwarning("Warning", "Excel file was not saved.")

if __name__ == "__main__":
    app = HandoverChecklist()
    app.mainloop()